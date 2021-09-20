
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border,Side
from openpyxl.styles.borders import *


wb1 = openpyxl.load_workbook('C:/Users/fun-f/Desktop/棚卸/【2021】 棚卸ロス集計表.xlsx')
season = input()


ws = wb1.active

#ws_1 = wb1["その２"]
ws_1 = wb1['確定 20210831']
print(ws_1)

#ws1 = wb1.worksheets[ws_1]

#print(ws1)

#ws_1.co

#if season == str(1) :
  #ws1.insert_cols(2,3)
  
#elif season == str(2):
  #ws1.insert_cols(5,3)
  
#elif season == str(3) :
  #ws1.insert_cols(8,3)  
  
if season == str(1):
  
  #1月は28列
  col_count = ws_1.max_column
  print(col_count)
  block = 2
  in_block = int(season) * block
  ws_1.insert_cols(col_count - 2,in_block)  
  
  title1 = str(season) + "月ロス率"

  title2 = str(season) + "月ロス率（絶対値)"
  
  this_cell1 = "X4"
  this_cell2 = "Y4"
  side = Side(style='thin', color='000000')
  border = Border(top=side, bottom=side, left=side, right=side)#t通常罫線
  
  border2 = Border(top=Side(style=BORDER_DOUBLE), bottom=Side(style=BORDER_DOUBLE), left=Side(style=BORDER_DOUBLE), right=Side(style=BORDER_THIN))
  
  border2_1 = Border(top=Side(style=BORDER_DOUBLE), bottom=Side(style=BORDER_THIN), left=Side(style=BORDER_DOUBLE), right=Side(style=BORDER_THIN))
  
  border2_2 = Border(top=Side(style=BORDER_THIN), bottom=Side(style=BORDER_DOUBLE), left=Side(style=BORDER_DOUBLE), right=Side(style=BORDER_THIN))
  
  border2_3 = Border(top=Side(style=BORDER_THIN), bottom=Side(style=BORDER_THIN), left=Side(style=BORDER_DOUBLE), right=Side(style=BORDER_THIN))
  
  border2_4 = Border(top=Side(style=BORDER_THIN), bottom=Side(style=BORDER_DOUBLE), left=Side(style=BORDER_DOUBLE), right=Side(style=BORDER_THIN))
  
  #-------------------------------------------
  
  border3 = Border(top=Side(style=BORDER_DOUBLE), bottom=Side(style=BORDER_DOUBLE), left=Side(style=BORDER_THIN), right=Side(style=BORDER_DOUBLE))  
  border3_1 = Border(top=Side(style=BORDER_DOUBLE), bottom=Side(style=BORDER_THIN), left=Side(style=BORDER_THIN), right=Side(style=BORDER_DOUBLE))   

  border3_2 = Border(top=Side(style=BORDER_THIN), bottom=Side(style=BORDER_DOUBLE), left=Side(style=BORDER_THIN), right=Side(style=BORDER_DOUBLE))  
  
  border3_3 = Border(top=Side(style=BORDER_THIN), bottom=Side(style=BORDER_THIN), left=Side(style=BORDER_THIN), right=Side(style=BORDER_DOUBLE))  
  
  border3_4 = Border(top=Side(style=BORDER_THIN), bottom=Side(style=BORDER_DOUBLE), left=Side(style=BORDER_THIN), right=Side(style=BORDER_DOUBLE))  
  
  ['BORDER_DASHDOT',
 'BORDER_DASHDOTDOT',
 'BORDER_DASHED',
 'BORDER_DOTTED',
 'BORDER_DOUBLE',
 'BORDER_HAIR',
 'BORDER_MEDIUM',
 'BORDER_MEDIUMDASHDOT',
 'BORDER_MEDIUMDASHDOTDOT',
 'BORDER_MEDIUMDASHED',#破線
 'BORDER_NONE',
 'BORDER_SLANTDASHDOT',
 'BORDER_THICK',
 'BORDER_THIN']#
  
  
  #★★★　X列の罫線編集　★★★
  
  row_count = ws_1.max_row
  upper_r = 5 #上段
  lower_r = 12 #下段
  turn_count = row_count - (upper_r + lower_r) - 1
  
  ws_1[this_cell1].value = title1
  ws_1[this_cell1].font = openpyxl.styles.fonts.Font(name="ＭＳ Ｐゴシック",size=11)
  ws_1[this_cell1].alignment = Alignment(horizontal='center',vertical='center')
  ws_1[this_cell1].fill = openpyxl.styles.PatternFill(patternType='solid',fgColor='D1FE7B',bgColor='D1FE7B')
  ws_1[this_cell1].border = border2
  
  ws_1["X" + str(upper_r + 1)].border = border2_1#最上段CELL
  ws_1["Y" + str(upper_r + 1)].border = border3_1#最上段CELL
  ws_1["X" + str(row_count - lower_r -1)].border = border2_4#最上段CELL
  ws_1["Y" + str(row_count - lower_r -1)].border = border3_4#最上段CELL  
  
  for i in range(turn_count):
    cell_i = upper_r + i + 1
    cell_c = "X" + str(cell_i)
    print(cell_c)
    cell_d = "Y" + str(cell_i)
    
    ws_1[cell_c].border = border2_3
    ws_1[cell_d].border = border3_3
  
  
  print(row_count)
  
  ws_1["X27"].border = border2
  ws_1["X28"].border = border2_1
  ws_1["X29"].border = border2_2
  ws_1["X30"].border = border2
  ws_1["X31"].border = border2
  ws_1["X32"].border = border2
  #ws_1.column_dimensions[col_no[9]].width = 12
  
  #★★★　X列の罫線編集　★★★
  ws_1["Y4"].value = title2
  ws_1[this_cell2].font = openpyxl.styles.fonts.Font(name="ＭＳ Ｐゴシック",size=11)
  ws_1[this_cell2].alignment = Alignment(horizontal='center',vertical='center')
  ws_1[this_cell2].fill = openpyxl.styles.PatternFill(patternType='solid',fgColor='D1FE7B')#,bgColor='D1FE7B'
  ws_1[this_cell2].border = border3
  ws_1["Y27"].border = border3
  ws_1["Y28"].border = border3_1
  ws_1["Y29"].border = border3_2
  ws_1["Y30"].border = border3
  ws_1["Y31"].border = border3
  ws_1["Y32"].border = border3
  
  for i in range(6,26):
    ws_1["X" + str(i)].value = "=T" + str(i)
    ws_1["X" + str(i)].border = border
    ws_1["Y" + str(i)].value = "=W" + str(i)
    ws_1["Y" + str(i)].border = border
  
elif season == str(2):
  #2月は30列
  block = 2
  in_block = int(season) * block
  ws_1.insert_cols(24,in_block)    
  
elif season == str(3):
  #3月は32列
  block = 3
  in_block = int(season) * block
  ws_1.insert_cols(24,in_block)    
    
elif season == str(3):
  #4月は34列
  block = 3
  in_block = int(season) * block
  ws_1.insert_cols(24,in_block)    
        
    
  
cols_count = ws_1.max_column
    
print(cols_count)   
  
#ws_1["F1"].value = ws1["A1"].value + ws1["D1"].value
#ws_1["G1"].value = "=(A1 + D1)"
#ws1.insert_rows(17,3)


wb1.save('C:/Users/fun-f/Desktop/棚卸/【2021】 棚卸ロス集計表.xlsx')
wb1.close()
print(ws_1)



  
  
